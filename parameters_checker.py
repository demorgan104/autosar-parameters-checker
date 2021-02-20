"""This module will help you determine if your 
   AUTOSAR project is properly configured by comparing the parameters in the configuration
   with the parameters specified in the requirements coming from a requirement 
   management system.

   Adapt it for your project if necessary and use it in CI as a verification method 
   for AUTOSAR code configuration for 3rd party code.

   It will print the analysis logs in the console. 
   No logger is required/implemented for the moment.

   The result of the analysis is stored by default in the out directory.
   To use this module, please make sure your input is similar to the one provided as a sample 
   in the input directory.
"""
import os
import re
import argparse
import shutil
import openpyxl


def find_arxml_files(search_path):
    """Generates a list of arxml files

    Args:
        search_path: The directory where to look for arxml files

    Yields:
        arxml_file_path
    """
    for subdir, _dirs, files in os.walk(search_path):
        for file in files:
            if file.endswith(".arxml"):
                yield os.path.join(subdir, file)


def search_for_param(parameters, search_path):
    """Search for parameters in arxml files at search_path

    Args:
        parameters: param_name:value pairs
        search_path: a path that points to a directory containing arxml files
    Returns:
        {
            param_name_1: [
                [
                    file_where_found,
                    line,
                    actual_value
                ],
                [
                    file_where_found,
                    line,
                    actual_value
                ]
            ],
            param_name_2: [
                [
                    file_where_found,
                    line,
                    actual_value
                ],
                [
                    file_where_found,
                    line,
                    actual_value
                ]
            ],
        }
    """
    param_data = {}
    value_pattern = re.compile(r"<VALUE>(.*)</VALUE>")
    param_pattern = re.compile(r">/(.*)</")
    for arxml_file in find_arxml_files(search_path):
        with open(arxml_file, "r", encoding="utf-8") as xml_file:
            print("analysing file {} ...".format(arxml_file))
            for line in xml_file:
                if "DEFINITION-REF" in line:
                    param_match = param_pattern.search(line)
                    if not param_match:
                        continue
                    param_name = param_match.group(1).split("/")[-1]
                    if param_name in parameters:
                        data_param = [arxml_file, line.strip()]
                        # next line should contain the value
                        next_line = next(xml_file)
                        match = value_pattern.match(next_line.strip())
                        if match:
                            value = match.group(1)
                            data_param.append(value)
                            if param_name in param_data:
                                param_data[param_name].append(data_param)
                            else:
                                param_data[param_name] = [data_param]
                        else:
                            print(
                                "Discarding param {} found at line  \n{}.   \nCould not find a value on next line in file {}".format(
                                    param_name, line, arxml_file
                                )
                            )
    return param_data


def compute_param(param_name, param_expected_value, param_result):
    """Computes a parameter value and modifies the container dicts with the computed value

    Args:
        param_name (string): Parameter name in the input file
        param_expected_value (string): Expected value extracted from the input file
        param_result (list): A list of results for the parameter after the arxml files were parsed
    Returns:
        A tuple of dictionaries containing the result of the computation
    """
    values_as_expected = {}
    cannot_decide_value = {}
    not_equal_expected_actual_values = {}
    found_field = ""
    actual_value_field = ""
    value_decider = set()
    for result in param_result:
        file = result[0]
        line = result[1]
        actual_value = result[2]
        found_field += "File: {file} at line: {line}   <br />".format(
            file=file, line=line
        )
        actual_value_field += (
            'Value in {file}: <span style="color:red">{val}</span>   <br />'.format(
                file=file, val=actual_value
            )
        )
        value_decider.add(actual_value)
    if len(value_decider) > 1:
        actual_value_field += "Could not decide the value. Multiple different values exists. Check above and correct"
        cannot_decide_value[param_name] = get_container_dict(
            param_name, param_expected_value, actual_value_field, found_field
        )
    else:
        computed_value = next(iter(value_decider))
        # Uppercase the values and get rid of measurement units not available in config
        processed_computed_value = computed_value.upper().split(" ")[0]
        processed_expected_value = param_expected_value.upper().split(" ")[0]
        actual_value_field += (
            'Computed actual value: <span style="color:green">{}</span>'.format(
                processed_computed_value
            )
        )
        if processed_computed_value == processed_expected_value:
            values_as_expected[param_name] = get_container_dict(
                param_name, param_expected_value, actual_value_field, found_field
            )
        else:
            not_equal_expected_actual_values[param_name] = get_container_dict(
                param_name, param_expected_value, actual_value_field, found_field
            )
    return values_as_expected, cannot_decide_value, not_equal_expected_actual_values


def get_container_dict(
    param_name, param_expected_value, actual_value_field, found_field
):
    """Returns a dictionary with pre-defined keys and parameters as values.
    """
    return {
        "name": param_name,
        "expected_value": param_expected_value,
        "actual_value": actual_value_field,
        "found_in": found_field,
    }


def compute_params(parameters, param_search_result):
    """Computes the parameters based on the search result in the arxml files

    Args:
        parameters: [description]
        param_search_result: [description]

    Returns:
        A tuple of dictionaries containing the computation result for parameters
    """
    values_as_expected = {}
    nothing_found_container = {}
    cannot_decide_value = {}
    not_equal_expected_actual_values = {}
    for param_name, param_expected_value in parameters.items():
        try:
            param_result = param_search_result[param_name]
        except KeyError:
            print("No results found for {}".format(param_name))
            nothing_found_container[param_name] = {
                "name": param_name,
                "expected_value": param_expected_value,
                "actual_value": "not found",
                "found_in": "not found",
            }
            continue
        result = compute_param(param_name, param_expected_value, param_result)
        values_as_expected = {**values_as_expected, **result[0]}
        cannot_decide_value = {**cannot_decide_value, **result[1]}
        not_equal_expected_actual_values = {**not_equal_expected_actual_values, **result[2]}

    return (
        values_as_expected,
        nothing_found_container,
        cannot_decide_value,
        not_equal_expected_actual_values,
    )


def write_report_header(report_file):
    """
    Creates a markdown table header
    """
    report_table_header = """NAME | EXPECTED VALUE | ACTUAL VALUE | FOUND IN |
--- | --- | --- | ---"""
    report_file.write(report_table_header + "\n")


def write_line(report_file, line_dict):
    """Inserts a markdown table line in report_file

    Args:
        report_file
        line_dict
    """
    line_template = "{name} | {expected_value} | {actual_value} | {found_in}\n"
    report_file.write(
        line_template.format(
            name=line_dict["name"],
            expected_value=line_dict["expected_value"],
            actual_value=line_dict["actual_value"],
            found_in=line_dict["found_in"],
        )
    )


def generate_report(excel_cells, computed_dicts, excel_file, out_dir):
    """Generates a report and updates the excel_file cells

    Args:
        excel_cells: Excel cells to be updated for each parameter
        computed_dicts: Parameters computation result
        excel_file: Excel file to save at the end of the update process
        out_dir: Save the reports here
    """
    with open(os.path.join(out_dir, "report.md"), "w+") as report_file:
        report_file.write("# PARAMETERS WITH VALUES AS EXPECTED\n")
        write_report_header(report_file)
        # Write the params where everything is ok
        for param_name, container_dict in computed_dicts[0].items():
            write_line(report_file, container_dict)
            excel_cells[param_name][3].value = "Agreed"
            excel_cells[param_name][6].value = "Found in config and value is matching."

        # Write the parameters that were not found
        report_file.write("# PARAMETERS THAT WERE NOT FOUND IN CONFIG\n")
        # Write the header for another table
        write_report_header(report_file)
        for param_name, container_dict in computed_dicts[1].items():
            write_line(report_file, container_dict)
            excel_cells[param_name][3].value = "Follow-up"
            excel_cells[param_name][6].value = "Not found in config"

        # Write the parameters where a value could not be decided
        report_file.write("\n# PARAMETERS WITH VALUES THAT COULD NOT BE COMPUTED\n")
        # Write the header for another table
        write_report_header(report_file)
        for param_name, container_dict in computed_dicts[2].items():
            write_line(report_file, container_dict)
            excel_cells[param_name][3].value = "Follow-up"
            excel_cells[param_name][6].value = "Multiple different values in config"
        # Write the parameters were values don't match
        report_file.write("\n# PARAMETERS WITH WRONG VALUE\n")
        # Write the header for another table
        write_report_header(report_file)
        for param_name, container_dict in computed_dicts[3].items():
            write_line(report_file, container_dict)
            excel_cells[param_name][3].value = "Follow-up"
            excel_cells[param_name][
                6
            ].value = "Values available in config but don't match"

        excel_file.save(os.path.join(out_dir, "new_report.xlsx"))


def get_param(line):
    """

    Args:
        line: A line of the form : <parameter_name> shall be set to <parameter_value>

    Returns:
        tuple: (
            <parameter_name>, 
            <parameter_value>
        )
    """
    pattern = re.compile("(.*)shall be set to.(.*[\\w])")
    match = pattern.match(line)
    if match:
        return (match.group(1), match.group(2))
    return None


def fill_params_dict(object_type, text, parameters):
    """
    Populates parameters with data

    Args:
        object_type: the object type should always be "Requirement"
        text: the text to match

    Returns:
        parameter name if successful
    """
    if "Requirement" in object_type:
        res = get_param(text.replace("\n", " "))
        if res:
            param_name = res[0].strip()
            param_value = res[1]
            if param_name in parameters:
                print(
                    "Parameter {} found multiple times. Considering only the first occurence.".format(
                        param_name
                    )
                )
            else:
                parameters[param_name] = param_value
                return param_name
        else:
            print("Could not match {}".format(text))
    else:
        pass
    return None


def extract_params_from_excel(excel_file):
    """Extracts a set of parameters from an input excel file

    The excel file columns are for the moment hard coded.

    Columns

    Returns:
        A tuple of the form : 
            Index 0: the parameters extracted
            Index 1: the cells for the parameters
            Index 2: the excel workbook
    """
    parameters = {}
    parameters_cells = {}
    excel_workbook = openpyxl.load_workbook(excel_file)
    active_sheet = excel_workbook.active

    for cell in active_sheet.iter_rows(
        min_col=2, max_col=8, min_row=1, max_row=active_sheet.max_row
    ):
        object_type = cell[0].value
        text = cell[1].value
        param_name = fill_params_dict(object_type, text, parameters)
        if param_name:
            parameters_cells[param_name] = cell

    print("Found {} parameters".format(len(parameters.keys())))
    # sys.exit(1)
    return parameters, parameters_cells, excel_workbook


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Find AUTOSAR configuration values.")
    parser.add_argument(
        "--config-location",
        help="Points to AUTOSAR config files directory",
    )
    parser.add_argument(
        "--input",
        help="Excel input file",
    )
    parser.add_argument(
        "--output-dir",
        help="Output directory",
        default="_out"
    )
    args = parser.parse_args()

    if os.path.exists(args.output_dir):
        shutil.rmtree(args.output_dir)
    
    os.mkdir(args.output_dir)

    sip_location = os.path.join(args.config_location)
    params, params_cells, excel = extract_params_from_excel(args.input)

    search_result = search_for_param(params, sip_location)

    computed_params = compute_params(params, search_result)

    generate_report(
        params_cells,
        computed_params,
        excel,
        args.output_dir
    )
